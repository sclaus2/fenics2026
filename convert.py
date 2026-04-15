from __future__ import annotations

import argparse
import csv
import difflib
import json
import re
import unicodedata
from collections import defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
from textwrap import dedent
from typing import Sequence

here = Path(__file__).parent

DEFAULT_BOOK_TITLE = "FEniCS Conference 2026 Book of Abstracts"
DEFAULT_BOOK_SUBTITLE = "Generated from conference abstract submissions"
DEFAULT_BOOK_AUTHOR = "FEniCS Conference 2026 Organizing Committee"
PLACEHOLDER_VALUES = {"na", "n/a", "none", "null", "nan", "-"}
TEXT_REPLACEMENTS = str.maketrans(
    {
        "\u00a0": " ",
        "\u2002": " ",
        "\u2003": " ",
        "\u2009": " ",
        "\u200a": " ",
        "\u202f": " ",
        "\u2010": "-",
        "\u2011": "-",
        "\u2012": "-",
        "\u2013": "-",
        "\u2014": "-",
        "\u2212": "-",
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u2026": "...",
        "\u2022": "*",
        "\u2192": "->",
    }
)
NAME_PREFIXES = ("dr", "dr.", "prof", "prof.", "professor")
SURNAME_PARTICLES = {
    "al",
    "ap",
    "ben",
    "bin",
    "da",
    "de",
    "del",
    "della",
    "der",
    "di",
    "dos",
    "du",
    "ibn",
    "la",
    "le",
    "mac",
    "mc",
    "st",
    "st.",
    "van",
    "von",
    "y",
}
SUBMISSION_TYPE_FIELD = (
    "Is the submission relating to a poster, a software demonstration or a presentation?\n\n"
    "Participants can demonstrate their FEniCS-based software to small groups using their laptops. "
    "These “software demonstrations” take place at the same time as the poster session. One "
    "participant can choose to submit a presentation or a poster and a software demonstration - "
    "please submit the form twice."
)
SUBMISSION_ORDER = {
    "Presentation": 0,
    "Poster": 1,
    "Software Demonstration": 2,
}

FIELD_ALIASES = {
    "email": ("Email address", "Username"),
    "presenter_name": ("Name of presenter",),
    "presenter_affiliation": ("Affiliation of presenter",),
    "submission_type": (SUBMISSION_TYPE_FIELD,),
    "title": ("Abstract title",),
    "text": ("Abstract text",),
    "authors": ("Name of authors (including presenter, comma-separated list)",),
    "affiliations": ("Affiliation of co-authors (including presenter, comma-separated list)",),
    "references": ("Reference list",),
}

ABSTRACT_TEMPLATE = dedent(
    """\
    ---
    title: {title}
    {authors}
    license: CC-BY-4.0
    exports:
      - format: pdf
        template: ../../template

    ---

    {metadata}

    {text}
    {references}

    """
)

README_TEMPLATE = dedent(
    """\
    ---
    title: {book_title_yaml}
    authors:
      - name: {book_author}
    license: CC-BY-4.0
    exports:
      - format: pdf
        template: ../template
    ---

    # {book_title}

    {book_subtitle}

    Total abstracts: **{num_abstracts}**

    {sections}
    """
)

ALL_ABSTRACTS_TEMPLATE = dedent(
    """\
    ---
    title: {book_title_yaml}
    authors:
      - name: {book_author}
    license: CC-BY-4.0
    exports:
      - format: pdf
        template: ../template
    ---

    # {book_title}

    {book_subtitle}

    Total abstracts: **{num_abstracts}**

    {sections}

    # Abstracts

    {abstracts}
    """
)


def quote_yaml(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def normalise_text(value: str) -> str:
    cleaned = unicodedata.normalize("NFKC", value.replace("\r\n", "\n"))
    return cleaned.translate(TEXT_REPLACEMENTS).strip()


def is_placeholder(value: str) -> bool:
    return normalise_space(value).casefold().strip(".") in PLACEHOLDER_VALUES


def normalise_space(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip(" ,;")


def split_outside_brackets(value: str, delimiter: str = ",") -> list[str]:
    parts: list[str] = []
    current: list[str] = []
    round_depth = 0
    square_depth = 0
    curly_depth = 0
    for char in value:
        if char == "(":
            round_depth += 1
        elif char == ")" and round_depth > 0:
            round_depth -= 1
        elif char == "[":
            square_depth += 1
        elif char == "]" and square_depth > 0:
            square_depth -= 1
        elif char == "{":
            curly_depth += 1
        elif char == "}" and curly_depth > 0:
            curly_depth -= 1

        if char == delimiter and round_depth == 0 and square_depth == 0 and curly_depth == 0:
            part = "".join(current).strip()
            if part:
                parts.append(part)
            current = []
            continue
        current.append(char)

    part = "".join(current).strip()
    if part:
        parts.append(part)
    return parts


def split_name_chunks(value: str) -> list[str]:
    if not value:
        return []

    chunks: list[str] = []
    for part in split_outside_brackets(value.replace("\n", " "), ","):
        for subpart in re.split(r"\s+(?:and|&)\s+", part):
            cleaned = normalise_space(subpart)
            if cleaned:
                chunks.append(cleaned)
    return chunks


def slugify(value: str) -> str:
    normalised = unicodedata.normalize("NFKD", value)
    ascii_value = normalised.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_value.lower()).strip("-")
    return slug or "abstract"


def escape_table(value: str) -> str:
    return value.replace("|", "\\|")


def escape_latex(value: str) -> str:
    replacements = {
        "\\": r"\textbackslash{}",
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
        "{": r"\{",
        "}": r"\}",
        "~": r"\textasciitilde{}",
        "^": r"\textasciicircum{}",
    }
    return "".join(replacements.get(char, char) for char in value)


def normalise_header(value: str) -> str:
    normalised = unicodedata.normalize("NFKC", value.replace("\r\n", "\n").replace("\r", "\n"))
    collapsed = re.sub(r"\s+", " ", normalised).strip()
    return collapsed.casefold()


def find_value(row: dict[str, str], field: str, default: str = "") -> str:
    normalised_row = {normalise_header(key): value for key, value in row.items()}
    for alias in FIELD_ALIASES[field]:
        if alias in row and row[alias].strip():
            value = normalise_text(row[alias])
            if is_placeholder(value):
                return default
            return value
        normalised_alias = normalise_header(alias)
        if normalised_alias in normalised_row and str(normalised_row[normalised_alias]).strip():
            value = normalise_text(str(normalised_row[normalised_alias]))
            if is_placeholder(value):
                return default
            return value
        truncated_matches = [
            key
            for key in normalised_row
            if len(key) >= 32 and (normalised_alias.startswith(key) or key.startswith(normalised_alias))
        ]
        if truncated_matches:
            best_key = max(truncated_matches, key=len)
            value = normalise_text(str(normalised_row[best_key]))
            if is_placeholder(value):
                return default
            return value
    return default


def extract_affiliation_markers(value: str) -> list[str]:
    markers: list[str] = []
    marker_groups = []
    for pattern in (
        r"\^\{([^}]*)\}",
        r"\^\(([^)]*)\)",
        r"\(([0-9,\s]+)\)\)?$",
    ):
        marker_groups.extend(re.findall(pattern, value))
    for group in marker_groups:
        markers.extend(re.findall(r"\d+", group))
    return list(dict.fromkeys(markers))


def has_affiliation_markers(value: str) -> bool:
    return bool(extract_affiliation_markers(value) or re.search(r"\[[0-9,\s]+\]$", value))


def clean_author_name(value: str) -> str:
    cleaned = re.sub(r"\^\{[^}]*\}", "", value)
    cleaned = re.sub(r"\^\([^)]*\)", "", cleaned)
    cleaned = re.sub(r"\(([0-9,\s]+)\)\)?$", "", cleaned)
    cleaned = re.sub(r"\[[0-9,\s]+\]$", "", cleaned)
    cleaned = normalise_space(cleaned).strip("()")
    parts = cleaned.split()
    while parts and parts[0].casefold().rstrip(".") in NAME_PREFIXES:
        parts.pop(0)
    return " ".join(parts)


def looks_like_person_name(value: str) -> bool:
    if not value:
        return False
    if any(char.isdigit() for char in value):
        return False

    tokens = re.findall(r"[A-Za-zÀ-ÿ][A-Za-zÀ-ÿ.'’-]*", value)
    return 1 <= len(tokens) <= 8


def parse_author_entries(value: str) -> list[tuple[str, list[str]]]:
    entries: list[tuple[str, list[str]]] = []
    for chunk in split_name_chunks(value):
        cleaned = clean_author_name(chunk)
        if looks_like_person_name(cleaned):
            entries.append((cleaned, extract_affiliation_markers(chunk)))
    return entries


def name_tokens(value: str) -> list[str]:
    normalised = unicodedata.normalize("NFKD", clean_author_name(value).casefold())
    ascii_value = normalised.encode("ascii", "ignore").decode("ascii")
    return re.findall(r"[a-z]+", ascii_value)


def sortable_name_text(value: str) -> str:
    tokens = name_tokens(value)
    return " ".join(tokens) if tokens else clean_author_name(value).casefold()


def presenter_sort_key(presenter: str) -> tuple[str, str, str]:
    cleaned = clean_author_name(presenter)
    if not cleaned:
        return ("", "", "")

    parts = cleaned.split()
    surname_parts = [parts[-1]]
    index = len(parts) - 2
    while index >= 0 and parts[index].casefold().rstrip(".") in SURNAME_PARTICLES:
        surname_parts.insert(0, parts[index])
        index -= 1

    surname_length = len(surname_parts)
    given_parts = parts[:-surname_length]
    surname = " ".join(surname_parts)
    given = " ".join(given_parts)
    return (
        sortable_name_text(surname),
        sortable_name_text(given),
        sortable_name_text(cleaned),
    )


def tokens_compatible(shorter: Sequence[str], longer: Sequence[str]) -> bool:
    if not shorter:
        return True
    remaining = list(longer)
    for token in shorter:
        match_index = next(
            (
                index
                for index, candidate in enumerate(remaining)
                if candidate == token
                or candidate.startswith(token)
                or token.startswith(candidate)
                or (len(token) == 1 and candidate.startswith(token))
                or (len(candidate) == 1 and token.startswith(candidate))
            ),
            None,
        )
        if match_index is None:
            return False
        remaining = remaining[match_index + 1 :]
    return True


def same_person_name(left: str, right: str) -> bool:
    left_tokens = name_tokens(left)
    right_tokens = name_tokens(right)
    if not left_tokens or not right_tokens:
        return False
    if left_tokens == right_tokens:
        return True
    if left_tokens[-1] != right_tokens[-1]:
        return False
    if left_tokens[0][0] != right_tokens[0][0]:
        return False

    left_middle = left_tokens[:-1]
    right_middle = right_tokens[:-1]
    shorter, longer = sorted((left_middle, right_middle), key=len)
    return tokens_compatible(shorter, longer)


def presenter_matches_author(presenter: str, author: str) -> bool:
    if same_person_name(presenter, author):
        return True

    presenter_tokens = name_tokens(presenter)
    author_tokens = name_tokens(author)
    if not presenter_tokens or not author_tokens:
        return False
    if presenter_tokens[0][0] != author_tokens[0][0]:
        return False

    shorter, longer = sorted((presenter_tokens, author_tokens), key=len)
    return tokens_compatible(shorter, longer)


def clean_affiliation_text(value: str) -> str:
    return normalise_space(value.replace("\n", " "))


def parse_numbered_affiliations(value: str) -> dict[str, str]:
    if not value:
        return {}
    matches = list(re.finditer(r"(?:(?<=^)|(?<=\n)|(?<=,)\s*)(\d+)[.)]\s*", value))
    if not matches:
        return {}

    affiliations: dict[str, str] = {}
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(value)
        affiliation = clean_affiliation_text(value[start:end]).strip(",")
        if affiliation:
            affiliations[match.group(1)] = affiliation
    return affiliations


def strip_leading_author_list(value: str) -> str:
    parts = split_outside_brackets(value, ",")
    if not parts:
        return value

    first = parts[0]
    entries = parse_author_entries(first)
    if entries:
        return value[len(first) :].lstrip(" ,")
    return value


def parse_extra_author_entries_from_affiliations(value: str) -> list[tuple[str, list[str]]]:
    if not value:
        return []

    match = re.search(r"(?:(?<=^)|(?<=\n)|(?<=,)\s*)\d+[\.)]\s*", value)
    candidate = value[: match.start()] if match else value
    if not has_affiliation_markers(candidate):
        return []

    return parse_author_entries(candidate)


def dedupe_preserve_order(values: Sequence[str]) -> list[str]:
    return list(dict.fromkeys(values))


def canonicalise_affiliation(value: str) -> str:
    normalised = unicodedata.normalize("NFKD", normalise_space(value).casefold())
    ascii_value = normalised.encode("ascii", "ignore").decode("ascii")
    ascii_value = re.sub(r"\bdept\b", "department", ascii_value)
    ascii_value = re.sub(r"\buniv\b", "university", ascii_value)
    ascii_value = re.sub(r"[^a-z0-9]+", "", ascii_value)
    return ascii_value


def same_affiliation(left: str, right: str) -> bool:
    left_key = canonicalise_affiliation(left)
    right_key = canonicalise_affiliation(right)
    if not left_key or not right_key:
        return False
    if left_key == right_key:
        return True

    similarity = difflib.SequenceMatcher(a=left_key, b=right_key).ratio()
    length_gap = abs(len(left_key) - len(right_key))
    return similarity >= 0.985 and length_gap <= 3


def dedupe_affiliations(values: Sequence[str]) -> list[str]:
    unique: list[str] = []
    for value in values:
        if not value:
            continue
        if any(same_affiliation(value, existing) for existing in unique):
            continue
        unique.append(value)
    return unique


def strip_references_heading(value: str) -> str:
    lines = value.splitlines()
    while lines and not lines[0].strip():
        lines.pop(0)
    if lines and normalise_space(lines[0]).casefold().rstrip(":") == "references":
        lines.pop(0)
    return "\n".join(lines).strip()


def sanitise_references_text(value: str) -> str:
    text = strip_references_heading(value)
    text = re.sub(
        r"https?://(?:dx\.)?doi\.org/([^\s)]+)",
        lambda match: f"DOI: {match.group(1)}",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(
        r"\bdoi:\s*(10\.[^\s)]+)",
        lambda match: f"DOI: {match.group(1)}",
        text,
        flags=re.IGNORECASE,
    )
    return text


def build_authors(
    presenter_name: str,
    email: str,
    raw_authors: str,
    raw_affiliations: str,
) -> list["Author"]:
    presenter_name = clean_author_name(presenter_name)
    author_entries = parse_author_entries(raw_authors)

    ordered_entries: list[tuple[str, list[str]]] = []
    presenter_index: int | None = None

    for name, markers in author_entries:
        existing_index = next(
            (index for index, (existing_name, _) in enumerate(ordered_entries) if same_person_name(existing_name, name)),
            None,
        )
        if existing_index is not None:
            if not ordered_entries[existing_index][1] and markers:
                ordered_entries[existing_index] = (ordered_entries[existing_index][0], markers)
            if presenter_name and presenter_matches_author(presenter_name, name):
                presenter_index = existing_index
            continue

        ordered_entries.append((name, markers))
        if presenter_name and presenter_matches_author(presenter_name, name):
            presenter_index = len(ordered_entries) - 1

    if not ordered_entries and presenter_name:
        ordered_entries.insert(0, (presenter_name, []))
        presenter_index = 0

    numbered_affiliations = parse_numbered_affiliations(raw_affiliations)
    cleaned_affiliation_field = clean_affiliation_text(raw_affiliations)
    simple_affiliations = [
        clean_affiliation_text(part)
        for part in split_outside_brackets(cleaned_affiliation_field, ",")
        if clean_affiliation_text(part)
    ]

    authors: list[Author] = []
    for index, (name, markers) in enumerate(ordered_entries):
        affiliations: list[str] = []
        if markers and numbered_affiliations:
            affiliations = dedupe_affiliations(
                [numbered_affiliations[marker] for marker in markers if marker in numbered_affiliations]
            )
        elif simple_affiliations and len(simple_affiliations) == len(ordered_entries):
            affiliations = [simple_affiliations[index]]
        elif cleaned_affiliation_field:
            affiliations = [cleaned_affiliation_field]
        else:
            affiliations = ["Affiliation unavailable"]

        authors.append(
            Author(
                name=name,
                affiliations=affiliations or ["Affiliation unavailable"],
                email=email if presenter_index == index else "",
                name_is_literal=True,
            )
        )

    return authors


@dataclass
class Author:
    name: str
    affiliations: list[str]
    email: str = ""
    name_is_literal: bool = True

    @property
    def affiliation(self) -> str:
        return "; ".join(self.affiliations)

    def to_myst(self) -> str:
        lines = ["- name:", f"    literal: {quote_yaml(self.name)}"]
        lines.append("  affiliations:")
        for affiliation in self.affiliations:
            lines.append(f"    - {quote_yaml(affiliation)}")
        if self.email:
            lines.append(f"  email: {quote_yaml(self.email)}")
        return "\n".join(lines)


@dataclass
class Submission:
    slug: str
    title: str
    presenter: str
    presenter_affiliation: str
    submission_type: str
    authors: list[Author]
    text: str
    references: str

    def to_markdown(self) -> str:
        authors_block = "authors:\n" + "\n".join(f"  {line}" for author in self.authors for line in author.to_myst().splitlines())

        body = self.to_body_markdown()

        return ABSTRACT_TEMPLATE.format(
            title=quote_yaml(self.title),
            authors=authors_block,
            metadata=body["metadata"],
            text=body["text"],
            references=body["references"],
        )

    def to_body_markdown(self) -> dict[str, str]:
        metadata_lines = []
        if self.submission_type:
            metadata_lines.append(f"**Submission type:** {self.submission_type}")
        presenter_line = self.presenter
        if self.presenter_affiliation:
            presenter_line = f"{presenter_line} ({self.presenter_affiliation})"
        if presenter_line.strip():
            metadata_lines.append(f"**Presenter:** {presenter_line}")
        metadata = "\n\n".join(metadata_lines)

        references = ""
        if self.references:
            ref_text = "\n\n".join(part for part in sanitise_references_text(self.references).split("\n") if part.strip())
            references = f"\n# References\n{ref_text}\n"

        return {
            "metadata": metadata,
            "text": self.text,
            "references": references,
        }


def build_submission(row: dict[str, str], used_slugs: set[str]) -> Submission:
    title = find_value(row, "title")
    presenter = find_value(row, "presenter_name")
    presenter_affiliation = find_value(row, "presenter_affiliation")
    submission_type = find_value(row, "submission_type", default="Presentation")
    email = find_value(row, "email")
    text = find_value(row, "text")
    references = find_value(row, "references")
    authors = build_authors(
        presenter_name=presenter,
        email=email,
        raw_authors=find_value(row, "authors"),
        raw_affiliations=find_value(row, "affiliations"),
    )

    slug_base = slugify(title)
    slug = slug_base
    suffix = 2
    while slug in used_slugs:
        slug = f"{slug_base}-{suffix}"
        suffix += 1
    used_slugs.add(slug)

    return Submission(
        slug=slug,
        title=title,
        presenter=presenter,
        presenter_affiliation=presenter_affiliation,
        submission_type=submission_type,
        authors=authors,
        text=text,
        references=references,
    )


def sort_key(submission: Submission) -> tuple[int, str, str, str, str]:
    return (
        SUBMISSION_ORDER.get(submission.submission_type, len(SUBMISSION_ORDER)),
        *presenter_sort_key(submission.presenter),
        submission.title.casefold(),
    )


def render_sections(submissions: list[Submission]) -> str:
    grouped: defaultdict[str, list[Submission]] = defaultdict(list)
    for submission in submissions:
        grouped[submission.submission_type].append(submission)

    sections = []
    ordered_types = sorted(grouped, key=lambda item: SUBMISSION_ORDER.get(item, len(SUBMISSION_ORDER)))
    for index, submission_type in enumerate(ordered_types):
        items = sorted(grouped[submission_type], key=lambda item: (*presenter_sort_key(item.presenter), item.title.casefold()))
        section_title = f"{submission_type}s"
        if submission_type == "Software Demonstration":
            section_title = "Software Demonstration Session"
        latex_lines: list[str] = []
        if index > 0:
            latex_lines.extend([r"\clearpage", ""])
        latex_lines.append(rf"\section*{{{escape_latex(section_title)} ({len(items)})}}")
        if submission_type == "Software Demonstration":
            latex_lines.extend(["", r"\noindent These abstracts belong to the live software demonstration session.\par"])
        latex_lines.extend(
            [
                "",
                r"\begin{longtable}{p{0.74\textwidth}p{0.22\textwidth}}",
                r"\textbf{Title} & \textbf{Presenter} \\",
                r"\hline",
                r"\endfirsthead",
                r"\textbf{Title} & \textbf{Presenter} \\",
                r"\hline",
                r"\endhead",
            ]
        )
        for item in items:
            latex_lines.extend(
                [
                    rf"\href{{abstracts/{item.slug}.md}}{{{escape_latex(item.title)}}} & {escape_latex(item.presenter)} \\",
                ]
            )
        latex_lines.append(r"\end{longtable}")
        sections.append("\n".join(["```{raw} latex", *latex_lines, "```"]))
    return "\n\n".join(sections)


def render_combined_sections(submissions: list[Submission]) -> str:
    target_map = {
        submission.slug: f"abstract-{submission.slug}"
        for submission in sorted(submissions, key=sort_key)
    }
    grouped: defaultdict[str, list[Submission]] = defaultdict(list)
    for submission in submissions:
        grouped[submission.submission_type].append(submission)

    sections = []
    ordered_types = sorted(grouped, key=lambda item: SUBMISSION_ORDER.get(item, len(SUBMISSION_ORDER)))
    for index, submission_type in enumerate(ordered_types):
        items = sorted(grouped[submission_type], key=lambda item: (*presenter_sort_key(item.presenter), item.title.casefold()))
        section_title = f"{submission_type}s"
        if submission_type == "Software Demonstration":
            section_title = "Software Demonstration Session"
        latex_lines: list[str] = []
        if index > 0:
            latex_lines.extend([r"\clearpage", ""])
        latex_lines.append(rf"\section*{{{escape_latex(section_title)} ({len(items)})}}")
        if submission_type == "Software Demonstration":
            latex_lines.extend(["", r"\noindent These abstracts belong to the live software demonstration session.\par"])
        latex_lines.extend(
            [
                "",
                r"\begin{longtable}{p{0.74\textwidth}p{0.22\textwidth}}",
                r"\textbf{Title} & \textbf{Presenter} \\",
                r"\hline",
                r"\endfirsthead",
                r"\textbf{Title} & \textbf{Presenter} \\",
                r"\hline",
                r"\endhead",
            ]
        )
        for item in items:
            latex_lines.extend(
                [
                    rf"\hyperref[{target_map[item.slug]}]{{{escape_latex(item.title)}}} & {escape_latex(item.presenter)} \\",
                ]
            )
        latex_lines.append(r"\end{longtable}")
        sections.append("\n".join(["```{raw} latex", *latex_lines, "```"]))
    return "\n\n".join(sections)


def render_combined_abstracts(submissions: list[Submission]) -> str:
    target_map = {
        submission.slug: f"abstract-{submission.slug}"
        for submission in sorted(submissions, key=sort_key)
    }
    grouped: defaultdict[str, list[Submission]] = defaultdict(list)
    for submission in submissions:
        grouped[submission.submission_type].append(submission)

    sections = []
    ordered_types = sorted(grouped, key=lambda item: SUBMISSION_ORDER.get(item, len(SUBMISSION_ORDER)))
    for submission_type in ordered_types:
        items = sorted(grouped[submission_type], key=lambda item: (*presenter_sort_key(item.presenter), item.title.casefold()))
        section_title = f"{submission_type}s"
        if submission_type == "Software Demonstration":
            section_title = "Software Demonstration Session"
        lines = [f"## {section_title}"]
        for item in items:
            body = item.to_body_markdown()
            references = body["references"].replace("\n# References\n", "\n#### References\n")
            lines.extend(
                [
                    "",
                    f"({target_map[item.slug]})=",
                    f"### {item.title}",
                    "",
                    body["metadata"],
                    "",
                    body["text"],
                    references,
                ]
            )
        sections.append("\n".join(lines))
    return "\n\n".join(sections)


def write_book_pages(
    submissions: list[Submission],
    book_dir: Path,
    book_title: str,
    book_subtitle: str,
    book_author: str,
) -> None:
    book_dir.mkdir(parents=True, exist_ok=True)
    sections = render_sections(submissions)
    (book_dir / "README.md").write_text(
        README_TEMPLATE.format(
            book_title_yaml=quote_yaml(book_title),
            book_title=book_title,
            book_subtitle=book_subtitle,
            book_author=quote_yaml(book_author),
            num_abstracts=len(submissions),
            sections=sections,
        ),
        encoding="utf-8",
    )
    (book_dir / "all_abstracts.md").write_text(
        ALL_ABSTRACTS_TEMPLATE.format(
            book_title_yaml=quote_yaml(book_title),
            book_title=book_title,
            book_subtitle=book_subtitle,
            book_author=quote_yaml(book_author),
            num_abstracts=len(submissions),
            sections=render_combined_sections(submissions),
            abstracts=render_combined_abstracts(submissions),
        ),
        encoding="utf-8",
    )


def load_rows(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))

    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("Reading .xlsx files requires `openpyxl`. Install dependencies with `python3 -m pip install .`.")
            raise SystemExit(1)

        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        data: list[dict[str, str]] = []
        for values in rows[1:]:
            row = {
                header: "" if value is None else str(value).strip()
                for header, value in zip(headers, values)
                if header
            }
            data.append(row)
        return data

    print(f"Unsupported input format: {path.suffix}. Use .csv or .xlsx.")
    raise SystemExit(1)


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Convert conference abstract CSV/XLSX exports into MyST markdown.")
    parser.add_argument("path", type=Path, help="CSV or XLSX export from the conference abstract submission form.")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Directory for individual abstract markdown files. Defaults to book/abstracts.",
    )
    parser.add_argument(
        "--book-dir",
        type=Path,
        default=here / "book",
        help="Book directory where README.md and all_abstracts.md will be generated.",
    )
    parser.add_argument(
        "--keep-existing",
        action="store_true",
        help="Keep existing markdown files in the abstract output directory instead of clearing it first.",
    )
    parser.add_argument("--book-title", default=DEFAULT_BOOK_TITLE)
    parser.add_argument("--book-subtitle", default=DEFAULT_BOOK_SUBTITLE)
    parser.add_argument("--book-author", default=DEFAULT_BOOK_AUTHOR)
    args = parser.parse_args(argv)

    csv_path = args.path
    if not csv_path.is_file():
        print(f"{csv_path} is not a file")
        return 1

    output_dir = args.output or args.book_dir / "abstracts"
    output_dir.mkdir(parents=True, exist_ok=True)

    if not args.keep_existing:
        for markdown_file in output_dir.glob("*.md"):
            markdown_file.unlink()
        manifest_path = output_dir / "manifest.json"
        if manifest_path.exists():
            manifest_path.unlink()

    used_slugs: set[str] = set()
    submissions: list[Submission] = []
    for row in load_rows(csv_path):
        title = find_value(row, "title")
        text = find_value(row, "text")
        if not title or not text or is_placeholder(title) or is_placeholder(text):
            continue
        submissions.append(build_submission(row, used_slugs))

    submissions.sort(key=sort_key)

    for submission in submissions:
        (output_dir / f"{submission.slug}.md").write_text(submission.to_markdown(), encoding="utf-8")

    manifest = [asdict(submission) for submission in submissions]
    (output_dir / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    write_book_pages(submissions, args.book_dir, args.book_title, args.book_subtitle, args.book_author)

    print(f"Wrote {len(submissions)} abstracts to {output_dir}")
    print(f"Wrote generated book pages to {args.book_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
