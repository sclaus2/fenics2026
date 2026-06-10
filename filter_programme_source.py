from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from convert import (
    Submission,
    build_submission,
    find_programme_match,
    find_programme_match_index,
    find_value,
    is_placeholder,
    parse_programme,
)

here = Path(__file__).parent
default_programme_path = here / "programme.md" if (here / "programme.md").is_file() else here.parent / "programme.md"


@dataclass
class SourceRow:
    row_number: int
    values: list[object]
    submission: Submission


def load_source_rows(path: Path) -> tuple[list[object], list[SourceRow]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        headers = [cell.value for cell in worksheet[1]]
        header_names = [str(value).strip() if value is not None else "" for value in headers]
        used_slugs: set[str] = set()
        rows: list[SourceRow] = []

        for row_number in range(2, worksheet.max_row + 1):
            values = [worksheet.cell(row_number, column).value for column in range(1, worksheet.max_column + 1)]
            row = {
                header: "" if value is None else str(value).strip()
                for header, value in zip(header_names, values)
                if header
            }
            title = find_value(row, "title")
            text = find_value(row, "text")
            if not title or not text or is_placeholder(title) or is_placeholder(text):
                continue
            rows.append(
                SourceRow(
                    row_number=row_number,
                    values=values,
                    submission=build_submission(row, used_slugs),
                )
            )
        return headers, rows
    finally:
        workbook.close()


def select_programme_rows(rows: list[SourceRow], programme: Path) -> tuple[list[SourceRow], list[str], list[str]]:
    programme_entries = parse_programme(programme)
    selected: list[SourceRow] = []
    unused = list(rows)
    duplicates: list[str] = []
    unmatched: list[str] = []

    for entry in programme_entries:
        unused_submissions = [row.submission for row in unused]
        match_index = find_programme_match_index(unused_submissions, entry, allow_software_blitz=False)
        if match_index is None:
            match_index = find_programme_match_index(unused_submissions, entry, allow_software_blitz=True)
        if match_index is None:
            selected_submissions = [row.submission for row in selected]
            duplicate = find_programme_match(selected_submissions, entry, allow_software_blitz=False)
            if duplicate is None:
                duplicate = find_programme_match(selected_submissions, entry, allow_software_blitz=True)
            if duplicate is not None:
                duplicates.append(f"{entry.presenter} | {entry.title}")
            else:
                unmatched.append(f"{entry.submission_type} | {entry.presenter} | {entry.title}")
            continue

        selected.append(unused.pop(match_index))

    return selected, duplicates, unmatched


def write_workbook(headers: list[object], rows: list[SourceRow], output: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Form responses 1"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row.values)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    worksheet.freeze_panes = "A2"
    if worksheet.max_row > 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 80)

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    workbook.close()


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Create the canonical abstract source workbook from rows referenced by the programme."
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=here.parent / "AbstractSubmissionsFEniCS2026_corrected.xlsx",
        help="Corrected abstract submission workbook.",
    )
    parser.add_argument(
        "--programme",
        type=Path,
        default=default_programme_path,
        help="Programme markdown file used as the inclusion list.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=here.parent / "AbstractProgrammeFEniCS2026.xlsx",
        help="Filtered programme source workbook to write.",
    )
    args = parser.parse_args(argv)

    if not args.source.is_file():
        print(f"Missing source workbook: {args.source}")
        return 1
    if not args.programme.is_file():
        print(f"Missing programme file: {args.programme}")
        return 1
    if args.source.resolve() == args.output.resolve():
        print("Output path must differ from the source workbook.")
        return 1

    headers, rows = load_source_rows(args.source)
    selected, duplicates, unmatched = select_programme_rows(rows, args.programme)
    if unmatched:
        print("Unmatched programme entries:")
        for item in unmatched:
            print(f" - {item}")
        return 1

    write_workbook(headers, selected, args.output)
    print(f"Source rows considered: {len(rows)}")
    print(f"Programme rows written: {len(selected)}")
    print(f"Duplicate programme references skipped: {len(duplicates)}")
    for duplicate in duplicates:
        print(f" - {duplicate}")
    print(f"Wrote: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
